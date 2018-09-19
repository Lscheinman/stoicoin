#!/usr/bin/env python3
# -*- coding: utf-8 -*
import time, os, json
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime, timedelta

debugging = False

class POLERmap():
    '''
    Provide the user with the ability to:
    1) Upload a file representing transactional and master data
    2) Choose headers from the file
    3a) Allow selection of headers and mapping to POLE entities
    3b) Create entities based on POLE attributes
    3c) Make suggestions on mapping each to POLE
    4) Make relations between POLE entities identified
    '''

    def __init__(self, folder, DB, processed, config, upload):

        self.config    = config
        self.folder    = folder
        self.DB        = DB
        self.upload    = upload
        self.processed = processed
        self.getMaps()
        self.vMap      = self.setvMap(None)
        self.views     = []
        self.files     = []
        self.columns   = []
        self.entities  = []
        self.relations = []

    def getMaps(self):

        self.Maps = []
        for f in os.listdir(self.config):
            if f[:8] == 'POLE_MAP':
                with open('%s%s' % (self.config, f)) as Map:
                    self.Maps.append(json.loads(Map.read()))

        return self.Maps

    def setvMap(self, mapname):

        if mapname == None:
            CLASSIFICATION = 'B1'

            LABEL = 'LABEL'
            P_GEN = 'P_GEN'
            P_FNAME = 'P_FNAME'
            P_LNAME = 'P_LNAME'
            P_DOB = 'P_DOB'
            P_POB = 'P_POB'
            P_ORIGIN = 'P_ORIGIN'
            P_LOGSOURCE = 'P_LOGSOURCE'
            O_TYPE = 'O_TYPE'
            O_CATEGORY = 'O_CATEGORY'
            O_DESC = 'O_DESC'
            O_CLASS1 = 'O_CLASS1'
            O_CLASS2 = 'O_CLASS2'
            O_CLASS3 = 'O_CLASS3'
            O_ORIGIN = 'O_ORIGIN'
            O_LOGSOURCE = 'O_LOGSOURCE'
            R_TYPE = 'R_TYPE'
            S_LABEL = 'S_LABEL'
            T_LABEL = 'T_LABEL'

            vMap = {}
            vMap['mapname']   = 'VP'
            vMap['relations'] = [{S_LABEL : 'pSubject', T_LABEL : 'oMosaic', R_TYPE : 'HasAttribute'}]
            vMap['entities']  = ([{LABEL : 'pSubject', P_GEN : 'Gender', P_FNAME : 'Forename', P_LNAME : 'Surname', P_DOB : 'Age', P_POB : 'Town',
                                   P_ORIGIN : 'PersonID', P_LOGSOURCE : CLASSIFICATION},
                                {LABEL : 'oMosaic', O_TYPE : 'MOSAIC', O_CATEGORY : 'MOSAIC Type', O_DESC : 'MOSAIC Property', O_CLASS1 : 'MOSAIC HH Composition',
                                 O_CLASS2 : 'MOSAIC Age Brand', O_CLASS3: 'MOSAIC No of Children', O_ORIGIN : 'MOSAIC Household Income', O_LOGSOURCE : CLASSIFICATION}]
                                 )
            vMap['rules'] = [{'VP_RISK' : vMap['entities'][0][P_DOB] + vMap['entities'][0][P_LOGSOURCE]}]

        else:
            for vMap in self.Maps:
                if vMap['mapname'] == mapname:
                    self.vMap = vMap
        return vMap


    def getFile(self, fname):
        '''
        STEP 1:
        Get a view from a single Excel file using OpenXLS.
        For each sheet copy the contents into a dictionary that can be referenced for POLE extraction
        Headers for example will need to be presented in step 2 of POLERmapping
        '''
        fileURL = self.processed + fname
        self.files.append(fileURL)
        wb = load_workbook(filename=fileURL, read_only=False)
        for ws in wb.sheetnames:
            view = {}
            Data = wb[ws]
            view['Name'] = ws
            view['Headers'] = [cell.value for cell in Data[1]]
            view['Values'] = []
            i = 0
            for row in Data:
                values = {'index' : i}
                for key, cell in zip(view['Headers'], row):
                    values[key] = cell.value
                view['Values'].append(values)
                i+=1
        self.views.append(view)
        return view, fileURL
    
    
    def uploadFile(self, uploadURL):
        '''
        STEP 1 as controlled from models:
        Get a view from a single Excel file using OpenXLS.
        For each sheet copy the contents into a dictionary that can be referenced for POLE extraction
        Headers for example will need to be presented in step 2 of POLERmapping
        '''
        XL = ['csv', 'lsx', 'xls']
        data = []
        headers = [{'dsource' : '', 'attribute' : '', 'sample' : []}]
        for file in os.listdir(uploadURL):
            TS = datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')
            print("[%s-PolerService]Found file %s" % (TS, file))            
            
            if file[len(file)-2:] == 'py':
                print("Python File")
            elif file[len(file)-3:] in XL:
                wb = load_workbook(filename=uploadURL+file, read_only=False)
                for ws in wb.sheetnames:
                    view = {}
                    Data = wb[ws]
                    view['Name'] = ws
                    view['Headers'] = [cell.value for cell in Data[1]]
                    for h in view['Headers']:
                        headers.append({'dsource' : ws,
                                        'attribute' : h,
                                        'sample' : []})
                    view['Values'] = []
                    i = 0
                    for row in Data:
                        values = {'index' : i}
                        for key, cell in zip(view['Headers'], row):
                            values[key] = cell.value
                            for h in headers:
                                if h['attribute'] == key and h['dsource'] == ws:
                                    if len(h['sample']) < 5:
                                        if cell.value not in h['sample'] and cell.value != key:
                                            h['sample'].append(cell.value)
                        view['Values'].append(values)
                        i+=1
                    data.append(view)
    
                #self.moveFile('%s%s' % (self.upload, file), '%s%s' % (self.processed, file))
        
        
        summary = 'Cols: %s\nRows: %s' % (len(headers), i)
        response = {'status' : 200,
                    'count' : len(headers),
                    'cols' : headers,
                    'rows' : i,
                    'summary' : summary,
                    'file' : file}

        TS = datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')
        print("[%s-PolerService]Response %s" % (TS, response))

        return response

    def moveFile(self, fFrom, fTo):
        print("Moving file from %s to %s" % (fFrom, fTo))
        os.rename(fFrom, fTo)

    def getProcessedFiles(self):

        for f in os.listdir(self.processed):
            if f not in self.files:
                self.files.append(f)

        return self.files

    def getFolderMerge(self):

        totalView = []
        headers   = {}
        counts    = []

        # Turn each table into a dataframe and merge into a single frame
        for fname in os.listdir(self.folder):
            TS = datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')
            TS = time.time()
            print("[%s]: start: %s" % (TS, fname))
            view = self.getFile(fname)

            print("Complete after %d seconds" % (time.time()-TS))
            # Analyze the headers by count and data type
            for h in list(view):
                if h in headers:
                    headers[h]+=1
                else:
                    headers[h] = 1
                    headers
            totalView.append(view)

        TS = time.time()
        totalView = pd.concat(totalView)
        print("Merged files after %d seconds" % (time.time()-TS))
        TS = time.time()
        desktop = "C:\\Users\\d063195\\Desktop\\"
        writer = pd.ExcelWriter(desktop + 'TotalView.xlsx')
        totalView.to_excel(writer)
        print("[%s]: Saved to excel after %d seconds" % (datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S'), time.time()-TS))


    def getColumns(self):
        '''
        STEP 2:
        For earch of the views in the list of views extracted from a file
        get the headers. Take the 2nd value to ensure it looks at data and not the headers
        '''
        headers = []
        for view in self.views:
            for col in list(view['Headers']):
                headers.append({'Name' : view['Name'], 'col' : col, 'dtype' : type(view['Values'][2][col])})

        self.columns = self.columns + headers
        return headers

    def mapAttribute(self, eName, eAttribute, eVal):

        for e in self.entities:
            if e['Name'] == eName:
                e[eAttribute] = eVal
                return e

    def mapRelationship(self, eSource, eTarget, Label):

        RelationDefinition = {'S_LABEL' : eSource, 'T_LABEL' : eTarget, 'R_TYPE' : Label}
        self.relations.append(RelationDefinition)
        return RelationDefinition

    def createEntity(self, etype, eName):

        if str(etype).lower() == 'p':
            etype = self.createPerson()
        elif str(etype).lower() == 'o':
            etype = self.createObject()
        elif str(etype).lower() == 'l':
            etype = self.createLocation()
        elif str(etype).lower() == 'e':
            etype =  self.createEvent()

        etype['LABEL'] = eName
        self.entities.append(etype)

        return etype

    def createPerson(self):
        return {'TYPE' : 'Person', 'P_FNAME' : '', 'P_LNAME' : '', 'P_DOB' : '', 'P_ORIGIN' : '', 'P_LOGSOURCE' : '', 'P_POB' : '', 'P_GEN' : '', 'DESC' : ''}
    def createObject(self):
        return {'TYPE' : 'Object', 'O_TYPE' : '', 'O_CATEGORY' : '', 'O_DESC' : '', 'O_CLASS1' : '', 'O_CLASS2' : '', 'O_CLASS3': '', 'O_ORIGIN' : '', 'O_LOGSOURCE' : ''}
    def createLocation(self):
        return {'TYPE' : 'Location', 'L_TYPE' : '', 'L_DESC' : '', 'L_XCOORD' : '', 'L_YCOORD' : '', 'L_ZCOORD' : '', 'L_CLASS1': '', 'L_ORIGIN' : '', 'L_LOGSOURCE' : ''}
    def createEvent(self):
        return {'TYPE' : 'Event', 'E_TYPE' : '', 'E_CATEGORY' : '', 'E_DESC' : '', 'E_LANG' : '', 'E_CLASS1' : '', 'E_TIME' : '', 'E_DATE' : '', 'E_DTG' : '', 'E_XCOORD' : '', 'E_YCOORD' : '', 'E_ORIGIN' : '', 'E_ORIGINREF' : '', 'E_LOGSOURCE' : ''}

    def createMap(self, mName):

        Map = {}
        Map['mapname'] = mName
        Map['relations'] = self.relations
        Map['entities'] = self.entities
        Map['GUID'] = self.DB.insertObject('POLER', 'Map',
                                           str(Map), len(Map['entities']),
                                           len(Map['relations']), mName,
                                           'A1', None, None)
        self.Maps.append(Map)
        self.relations = []
        self.entities  = []

        return Map

    def POLERizeView(view):

        entities = []

        # Assign each column to a POLE entity and attribute
        for col in list(view):
            print(col)
            # Pick an entity that this column should belong too
            etype = input('Which type entity should %s belong to (POLE or existing)?\n%s' % (col, entities))
            print(etype)
            if str(etype).lower() == 'p':
                etype = {'P_FNAME' : '', 'P_LNAME' : '', 'P_DOB' : '', 'P_ORIGIN' : '', 'P_LOGSOURCE' : ''}
            elif str(etype).lower() == 'o':
                etype = {'O_TYPE' : '', 'O_CATEGORY' : '', 'O_DESC' : '', 'O_CLASS1' : '', 'O_CLASS2' : '', 'O_CLASS3': '', 'O_ORIGIN' : '', 'O_LOGSOURCE' : ''}
            elif str(etype).lower() == 'l':
                etype = {'L_TYPE' : '', 'L_DESC' : '', 'L_XCOORD' : '', 'L_YCOORD' : '', 'L_ZCOORD' : '', 'L_CLASS1': '', 'L_ORIGIN' : '', 'L_LOGSOURCE' : ''}
            elif str(etype).lower() == 'e':
                etype =  {'E_TYPE' : '', 'E_CATEGORY' : '', 'E_DESC' : '', 'E_LANG' : '', 'E_CLASS1' : '', 'E_TIME' : '', 'E_DATE' : '', 'E_DTG' : '', 'E_XCOORD' : '', 'E_YCOORD' : '', 'E_ORIGIN' : '', 'E_ORIGINREF' : '', 'E_LOGSOURCE' : ''}
            else:
                etype = entities.index(entities[etype])
                entities.remove(entities.index(entities[etype]))
            eatt = input('Which attribute should %s go to?\n%s' % (col, etype))
            etype[eatt] = col
            while str(input('Change another attribute?')).lower() != 'n':
                eatt = input('Which attribute?\n%s' % etype)
                eVal = input('Value?\n')
                etype[eatt] = eVal

            entities.append(etype)

        return entities

    def POLERExtract(self, view, file, fileGUID):

        '''
        For each row
           For each col in row.keys                check each column to see if it is part of an entity map
              For each entity in entities          cycle through each entity to compare keys to the row's current col
                 For each key in entity            for each key in the entity
                    if entity[key] == col          if the entity[key].value == the current col value
                       for nE in newEntities:      cycle through all the entities created so far
                          NewEntity = {}           if the current label is equal to the found entity label use this to map the new
                          NewEntity['LABEL'] = e   else create a new entity based on the LABEL first letter
                          NE[key] = col.value

        '''
        data = {'Persons' : [], 'Objects' : [], 'Events' : [], 'Locations' : []}
        for row in view['Values']:
            if row['index'] != 0:
                newE = []
                for c in row.keys():
                    for e in self.vMap['entities']:
                        for k in e.keys():
                            # The keys match so find the entity in the row
                            if e[k].strip().lower() == c.strip().lower():
                                found = False
                                for nE in newE:
                                    if found == False:
                                        if nE['LABEL'] == e['LABEL']:
                                            if e['LABEL'][0] == 'p':
                                                nE[k] = row[c]
                                                found = True
                                            elif e['LABEL'][0] == 'o':
                                                nE[k] = row[c]
                                                found = True
                                            elif e['LABEL'][0] == 'l':
                                                nE[k] = row[c]
                                                found = True
                                            else:
                                                nE[k] = row[c]
                                                found = True

                                # No existing entities so start with a new shell
                                if found == False:
                                    if e['LABEL'][0] == 'p':
                                        mE = self.createPerson()
                                    elif e['LABEL'][0] == 'o':
                                        mE = self.createObject()
                                    elif e['LABEL'][0] == 'l':
                                        mE = self.createLocation()
                                    else:
                                        mE = self.createEvent()
                                    mE['LABEL'] = e['LABEL']
                                    newE.append(mE)

                                # Assign the key value to this entity
                                mE[k] = row[c]

                # Check the entities extracted in the row and fill the GUID to complete the entity
                for mE in newE:
                    print("\n\n!!!!!\n%s" % mE)
                    if mE['TYPE'][0] == 'Person':
                        if mE not in data['Persons']:
                            mE['GUID'] = self.DB.insertPerson(mE['P_GEN'], mE['P_FNAME'], mE['P_LNAME'], mE['P_DOB'], mE['P_POB'], mE['P_ORIGIN'], None, mE['P_LOGSOURCE'], mE['DESC'])
                            data['Persons'].append(mE)

                    elif mE['TYPE'][0] == 'Object':
                        if mE not in data['Objects']:
                            mE['GUID'] = self.DB.insertObject(mE['O_TYPE'], mE['O_CATEGORY'], mE['O_DESC'], mE['O_CLASS1'], mE['O_CLASS2'], mE['O_CLASS3'], mE['O_ORIGIN'], None, mE['O_LOGSOURCE'])
                            data['Objects'].append(mE)

                    elif mE['TYPE'][0] == 'Location':
                        if mE not in data['Locations']:
                            mE['GUID'] = self.DB.insertLocation(mE['L_TYPE'], mE['L_DESC'], mE['L_XCOORD'], mE['L_YCOORD'], mE['L_ZCOORD'], mE['L_CLASS1'], mE['L_ORIGIN'], None, mE['L_LOGSOURCE'])
                            data['Locations'].append(mE)

                    else:
                        if mE not in data['Events']:
                            mE['GUID'] = self.DB.insertEvent(mE['E_TYPE'], mE['E_CATEGORY'], mE['E_DESC'], mE['E_LANG'], mE['E_CLASS1'], mE['E_TIME'], mE['E_DATE'], mE['E_DTG'], mE['E_XCOORD'], mE['E_YCOORD'], mE['E_ORIGIN'], None, mE['E_LOGSOURCE'])
                            data['Events'].append(mE)

                    self.DB.insertRelation(mE['GUID'], mE['TYPE'], 'FromFile', fileGUID, 'Object')

                # Use the relations table to determine the final steps
                for r in self.vMap['relations']:
                    for S in newE:
                        for T in newE:
                            if S['LABEL'] == r['S_LABEL'] and T['LABEL'] == r['T_LABEL']:
                                print("Insert rel %d %s %s %d %s" % ( S['GUID'], S['TYPE'], r['R_TYPE'], T['GUID'], T['TYPE']))

        self.moveFile('%s%s' % (self.processed, file), '%s%s' % (self.folder, file))


#TODO
# Map to Index through views and models
# Persist maps
#import OrientModels as om
#DB = om.OrientModel(None)
#folder = 'C:\\Users\\d063195\\Desktop\\_Projects\\20170120_VP\\Barnsley MBC Secure collaboration\\Vulnerable People PoC - Data files\\'
#fname = 'Person.xlsx'
#P = POLERmap(folder, DB)
#P.getFile(fname)
#headers = P.getColumns()
#P.POLERExtract(v)
'''
TODO
Full test of extraction without pandas


'''
